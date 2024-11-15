# -*- coding: utf-8 -*-
import threading
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service  # تم إضافة هذا السطر
from webdriver_manager.chrome import ChromeDriverManager
import time
import openpyxl
from concurrent.futures import ThreadPoolExecutor, as_completed
import re  # لإضافة التحقق من الروابط
import math  # لاستخدام الدالة ceil
import datetime  # لاستيراد datetime لطباعة الوقت


# إعداد خيارات المتصفح
chrome_options = Options()
chrome_options.add_argument('--headless')  # تشغيل Chrome في وضع الرأسية
chrome_options.add_argument('--no-sandbox')  # تجاوز نموذج الأمان في نظام التشغيل
chrome_options.add_argument('--disable-dev-shm-usage')  # استخدام /tmp بدلاً من /dev/shm

# إعداد خدمة ChromeDriver
service = Service(ChromeDriverManager().install())

# تهيئة المتصفح
driver = webdriver.Chrome(service=service, options=chrome_options)

# طباعة الوقت عند بدء السكربت
start_time = datetime.datetime.now()
print(f"The script started running at {start_time.strftime('%Y-%m-%d %H:%M:%S')}")

# قراءة الروابط من ملف Excel 'a3tmad_urls.xlsx' (تأكد من تغيير اسم الملف إذا لزم الأمر)
urls = []
try:
    urls_workbook = openpyxl.load_workbook('a3tmad_urls.xlsx')
    urls_sheet = urls_workbook.active
    for row in urls_sheet.iter_rows(values_only=True):
        url = row[0]  # نفترض أن الروابط موجودة في العمود الأول
        if url:
            url = url.strip()  # إزالة المسافات الزائدة
            if re.match(r'^https?://', url):  # التحقق من أن الرابط يبدأ بـ http:// أو https://
                urls.append(url)
            else:
                print(f"Invalid link ignored: {url}")
except Exception as e:
    print(f"An error occurred while reading the links file: {e}")
    exit()

# تحديد عدد الخيوط (يمكنك تعديل هذا الرقم بناءً على موارد جهازك)
max_workers = 10  # يمكنك تقليله إذا واجهت مشاكل في الأداء

# تحديد حجم الدفعة (عدد الروابط في كل دفعة)
batch_size = 10  # حفظ كل 1000 سجل في ملف واحد

# حساب عدد الدفعات
num_batches = math.ceil(len(urls) / batch_size)

# متغير لعدّ الملفات
file_counter = 1

# قائمة لحفظ الروابط التي لم يتم حفظها بسبب الأخطاء
error_links = []
error_links_lock = threading.Lock()

# متغير لحفظ الوقت المستغرق في الدفعة الأولى
first_batch_duration = None

# بدء معالجة الدفعات
for batch_num in range(num_batches):
    # إضافة الوقت عند بداية كل دفعة
    batch_start_time = datetime.datetime.now()
    print(f"Processing batch {batch_num + 1} of {num_batches} at {batch_start_time.strftime('%Y-%m-%d %H:%M:%S')}")
    # استخراج الدفعة الحالية من الروابط
    start_index = batch_num * batch_size
    end_index = min(start_index + batch_size, len(urls))
    current_batch_urls = urls[start_index:end_index]

    # إذا كانت هذه ليست الدفعة الأولى، قم بتقدير الوقت المتوقع للانتهاء
    if first_batch_duration and batch_num > 0:
        estimated_time = first_batch_duration
        estimated_completion_time = batch_start_time + datetime.timedelta(seconds=estimated_time)
        print(f"Estimated completion time for this batch: {estimated_completion_time.strftime('%Y-%m-%d %H:%M:%S')}")

    # إنشاء ملف Excel جديد لهذه الدفعة
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Projects"

    # كتابة عناوين الأعمدة (تمت إضافة "الرابط" كعمود جديد)
    sheet.append([
        "الرابط",  # العمود الجديد
        "اسم المنافسة",
        "رقم المنافسة",
        "الرقم المرجعي",
        "غرض المنافسة",
        "قيمة وثائق المنافسة",
        "حالة المنافسة",
        "هل التأمين من متطلبات المنافسة",
        "نوع المنافسة",
        "الجهة الحكومية",
        "طريقة تقديم العروض",
        "مطلوب ضمان الابتدائي",
        "تاريخ فحص العروض",
        "التاريخ المتوقع للترسية",
        "تاريخ بدء الأعمال / الخدمات",
        "مكان فتح العرض",
        "مكان التنفيذ",
        "الضمان النهائي",
        "نوع الاتفاقية",
        "مدة الاتفاقية",
        "التفاصيل",
        "تشمل المنافسة على بنود توريد",
        "اسم الموردين",
        "قيمة العرض المالي",
        "نتائج فحص العروض الفنية",
        "المورد المرسى عليه",
        "قيمة العرض للمورد المرسى عليه",
        "قيمة الترسية"
    ])

    # اسم الملف لحفظ الدفعة الحالية
    filename = f"a3tmad{file_counter}.xlsx"

    # قفل لحماية الوصول إلى الملف
    write_lock = threading.Lock()

    # دالة لاستخراج البيانات من رابط واحد
    def extract_data(url):
        driver = None  # تعريف المتغير بقيمة None
        try:
            # استخدام كائن Service مع ChromeDriverManager
            service = Service(ChromeDriverManager().install())
            driver = webdriver.Chrome(service=service, options=chrome_options)
            driver.get(url)
            time.sleep(1)  # الانتظار حتى يتم تحميل الصفحة بالكامل

            # استخراج البيانات المطلوبة مع التعامل مع البيانات الفارغة
            def get_text(xpath):
                try:
                    return driver.find_element(By.XPATH, xpath).text
                except:
                    return 'لايوجد'

            # استخراج البيانات
            اسم_المنافسة = get_text('//*[@id="basicDetials"]/div[2]/ul/li[1]/div/div[2]/span')
            رقم_المنافسة = get_text('//*[@id="basicDetials"]/div[2]/ul/li[2]/div/div[2]/span')
            الرقم_المرجعي = get_text('//*[@id="basicDetials"]/div[2]/ul/li[3]/div/div[2]/span')

            try:
                عنصر_النقر = driver.find_element(By.XPATH, '//*[@id="subPurposSapn"]/i')
                driver.execute_script("arguments[0].click();", عنصر_النقر)
                time.sleep(2)
            except:
                pass

            # تعديل استخراج الغرض من المنافسة لاستبعاد النص داخل <i>
            try:
                الغرض_من_المنافسة = driver.execute_script("""
                    var elem = document.getElementById('purposeSpan');
                    var text = '';
                    for (var i = 0; i < elem.childNodes.length; i++) {
                        if (elem.childNodes[i].nodeType === Node.TEXT_NODE) {
                            text += elem.childNodes[i].textContent;
                        }
                    }
                    return text.trim();
                """)
            except:
                الغرض_من_المنافسة = 'لايوجد'

            قيمة_وثائق_المنافسة = get_text('//*[@id="basicDetials"]/div[2]/ul/li[5]/div/div[2]/span')
            حالة_المنافسة = get_text('//*[@id="basicDetials"]/div[2]/ul/li[6]/div/div[2]/span')
            هل_التأمين_من_المتطلبات = get_text('//*[@id="basicDetials"]/div[2]/ul/li[8]/div/div[2]/span')
            نوع_المنافسة = get_text('//*[@id="basicDetials"]/div[3]/ul/li[1]/div/div[2]/span')
            الجهة_الحكومية = get_text('//*[@id="basicDetials"]/div[3]/ul/li[2]/div/div[2]/span')
            طريقة_تقديم_العروض = get_text('//*[@id="basicDetials"]/div[3]/ul/li[4]/div/div[2]/span')
            مطلوب_ضمان_الابتدائي = get_text('//*[@id="basicDetials"]/div[3]/ul/li[5]/div/div[2]/span')
            نوع_الاتفاقية = get_text('//*[@id="basicDetials"]/div[3]/ul/li[5]/div/div[2]/span')
            مدة_الاتفاقية = get_text('//*[@id="basicDetials"]/div[3]/ul/li[6]/div/div[2]/span')
            الضمان_النهائي = get_text('//*[@id="basicDetials"]/div[3]/ul/li[7]/div/div[2]/span')

            # العناوين والمواعيد المتعلقة بالمنافسة
            try:
                عنصر_النقر = driver.find_element(By.XPATH, '//*[@id="tenderDatesTab"]')
                driver.execute_script("arguments[0].click();", عنصر_النقر)
                time.sleep(1)
            except:
                pass

            تاريخ_فحص_العروض = get_text('//*[@id="offerDetials"]/div[2]/ul[1]/li[4]/div/div[2]/span[1]')
            التاريخ_المتوقع_للترسية = get_text('//*[@id="offerDetials"]/div[2]/ul[2]/li[1]/div/div[2]/span[1]')
            تاريخ_بدء_الاعمال_الخدمات = get_text('//*[@id="offerDetials"]/div[2]/ul[2]/li[2]/div/div[2]/span[1]')
            مكان_فتح_العرض = get_text('//*[@id="offerDetials"]/div[3]/ul/li/div/div[2]/span')

            # مجال التصنيف وموقع التنفيذ والتقديم
            try:
                عنصر_النقر = driver.find_element(By.XPATH, '//*[@id="relationStepTab"]')
                driver.execute_script("arguments[0].click();", عنصر_النقر)
                time.sleep(1)
            except:
                pass

            مكان_التنفيذ = get_text('//*[@id="ActivityDetials"]/div/ul[1]/li[1]/div/div[2]/div/div/ol/li/div/div/ul/li')
            التفاصيل = get_text('//*[@id="ActivityDetials"]/div/ul[1]/li[2]/div/div[2]/span')
            تشمل_المنافسة_علي_بنود_توريد = get_text('//*[@id="ActivityDetials"]/div/ul[2]/li[2]/div/div[2]/span')

            # النقر على العنصر لعرض أسماء الموردين
            try:
                عنصر_النقر = driver.find_element(By.XPATH, '//*[@id="awardingStepTab"]')
                driver.execute_script("arguments[0].click();", عنصر_النقر)
                time.sleep(2)
            except:
                pass

            # استخراج أسماء الموردين
            اسماء_الموردين = []
            for i in range(1, 60):
                المورد_xpath = f'//*[@id="offerDetials"]/div/table/tbody/tr[{i}]/td[1]'
                اسم_المورد = get_text(المورد_xpath)
                if اسم_المورد != 'لايوجد':
                    اسماء_الموردين.append(اسم_المورد)
                else:
                    break

            # استخراج قائمة قيمة العرض المالي
            قيمة_العرض_المالي_list = []
            for i in range(1, 60):
                قيمة_العرض_المالي_xpath = f'//*[@id="offerDetials"]/div/table/tbody/tr[{i}]/td[2]/h5'
                قيمة_العرض_المالي_value = get_text(قيمة_العرض_المالي_xpath)
                if قيمة_العرض_المالي_value != 'لايوجد':
                    قيمة_العرض_المالي_list.append(قيمة_العرض_المالي_value)
                else:
                    break

            # استخراج قائمة نتائج فحص العروض الفنية
            نتائج_فحص_العروض_الفنية_list = []
            for i in range(1, 60):
                نتائج_فحص_العروض_الفنية_xpath = f'//*[@id="offerDetials"]/div/table/tbody/tr[{i}]/td[3]/h5'
                نتائج_فحص_العروض_الفنية_value = get_text(نتائج_فحص_العروض_الفنية_xpath)
                if نتائج_فحص_العروض_الفنية_value != 'لايوجد':
                    نتائج_فحص_العروض_الفنية_list.append(نتائج_فحص_العروض_الفنية_value)
                else:
                    break

            # استخراج المورد المرسى عليه وقيم العروض
            المورد_المرسى_عليه = get_text('//*[@id="awardingDiv"]/div[2]/div/table/tbody/tr/td[1]')
            قيمة_العرض_المالي_للمورد = get_text('//*[@id="awardingDiv"]/div[2]/div/table/tbody/tr/td[2]/h5')
            قيمة_الترسية = get_text('//*[@id="awardingDiv"]/div[2]/div/table/tbody/tr/td[3]/h5')

            driver.quit()

            # جمع البيانات في قائمة
            rows_to_write = []
            if اسماء_الموردين:
                for idx, اسم_المورد in enumerate(اسماء_الموردين):
                    # الحصول على القيم المقابلة من القوائم الأخرى
                    قيمة_العرض_المالي = قيمة_العرض_المالي_list[idx] if idx < len(قيمة_العرض_المالي_list) else 'لايوجد'
                    نتائج_فحص_العروض_الفنية = نتائج_فحص_العروض_الفنية_list[idx] if idx < len(نتائج_فحص_العروض_الفنية_list) else 'لايوجد'

                    row = [
                        url,  # إضافة الرابط إلى البيانات
                        اسم_المنافسة,
                        رقم_المنافسة,
                        الرقم_المرجعي,
                        الغرض_من_المنافسة,
                        قيمة_وثائق_المنافسة,
                        حالة_المنافسة,
                        هل_التأمين_من_المتطلبات,
                        نوع_المنافسة,
                        الجهة_الحكومية,
                        طريقة_تقديم_العروض,
                        مطلوب_ضمان_الابتدائي,
                        تاريخ_فحص_العروض,
                        التاريخ_المتوقع_للترسية,
                        تاريخ_بدء_الاعمال_الخدمات,
                        مكان_فتح_العرض,
                        مكان_التنفيذ,
                        الضمان_النهائي,
                        نوع_الاتفاقية,
                        مدة_الاتفاقية,
                        التفاصيل,
                        تشمل_المنافسة_علي_بنود_توريد,
                        اسم_المورد,
                        قيمة_العرض_المالي,
                        نتائج_فحص_العروض_الفنية,
                        المورد_المرسى_عليه,
                        قيمة_العرض_المالي_للمورد,
                        قيمة_الترسية
                    ]
                    rows_to_write.append(row)
            else:
                row = [
                    url,  # إضافة الرابط إلى البيانات
                    اسم_المنافسة,
                    رقم_المنافسة,
                    الرقم_المرجعي,
                    الغرض_من_المنافسة,
                    قيمة_وثائق_المنافسة,
                    حالة_المنافسة,
                    هل_التأمين_من_المتطلبات,
                    نوع_المنافسة,
                    الجهة_الحكومية,
                    طريقة_تقديم_العروض,
                    مطلوب_ضمان_الابتدائي,
                    تاريخ_فحص_العروض,
                    التاريخ_المتوقع_للترسية,
                    تاريخ_بدء_الاعمال_الخدمات,
                    مكان_فتح_العرض,
                    مكان_التنفيذ,
                    الضمان_النهائي,
                    نوع_الاتفاقية,
                    مدة_الاتفاقية,
                    التفاصيل,
                    تشمل_المنافسة_علي_بنود_توريد,
                    'لايوجد',  # اسم الموردين
                    'لايوجد',  # قيمة العرض المالي
                    'لايوجد',  # نتائج فحص العروض الفنية
                    المورد_المرسى_عليه,
                    قيمة_العرض_المالي_للمورد,
                    قيمة_الترسية
                ]
                rows_to_write.append(row)

            # كتابة البيانات إلى الملف مع استخدام القفل
            with write_lock:
                for row in rows_to_write:
                    sheet.append(row)
                # حفظ الملف بعد كل كتابة
                workbook.save(filename)

        except Exception as e:
            print(f"An error occurred while processing the link: {url}")
            print(f"Error: {e}")
            if driver is not None:
                driver.quit()
            # إضافة الرابط إلى قائمة الروابط التي حدثت بها أخطاء
            with error_links_lock:
                error_links.append(url)

    # بدء الوقت لقياس مدة الدفعة
    batch_processing_start = time.time()

    # استخدام ThreadPoolExecutor لمعالجة الروابط في الدفعة الحالية بشكل متوازٍ
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = [executor.submit(extract_data, url) for url in current_batch_urls]

        # يمكنك استخدام as_completed لمتابعة تقدم المهام
        for future in as_completed(futures):
            try:
                future.result()
            except Exception as e:
                print(f"An error occurred while processing link: {e}")

    # حساب مدة معالجة الدفعة
    batch_processing_end = time.time()
    batch_duration = batch_processing_end - batch_processing_start

    # حساب الأيام والساعات والدقائق والثواني
    days = int(batch_duration // 86400)
    hours = int((batch_duration % 86400) // 3600)
    minutes = int((batch_duration % 3600) // 60)
    seconds = int(batch_duration % 60)
    print(f"The batch {batch_num + 1} was processed in {days} day(s), {hours} hour(s), {minutes} minute(s), and {seconds} second(s).")

    # إذا كانت هذه هي الدفعة الأولى، احفظ مدتها
    if batch_num == 0:
        first_batch_duration = batch_duration

    # زيادة عداد الملفات
    file_counter += 1

# بعد الانتهاء من جميع الدفعات، كتابة الروابط التي حدثت بها أخطاء إلى ملف Excel
if error_links:
    # إنشاء ملف Excel جديد للروابط التي حدثت بها أخطاء
    error_workbook = openpyxl.Workbook()
    error_sheet = error_workbook.active
    error_sheet.title = "Error Links"

    # كتابة عناوين الأعمدة
    error_sheet.append(["الروابط التي حدثت بها أخطاء"])

    # كتابة الروابط إلى الملف
    for link in error_links:
        error_sheet.append([link])

    # حفظ ملف Excel
    error_workbook.save("errors_links.xlsx")
    print("Error links have been saved to 'errors_links.xlsx'.")

print('All batches have been processed.')
